from openpyxl import load_workbook
from datetime import date
import requests
import json

URL_NIPY = 'https://wl-api.mf.gov.pl/api/search/nips/'

wb = load_workbook(filename='konta_bankowe.xlsx')
ws = wb['konta']

lista_kont = []
lista_plikow = []

url = URL_NIPY

for row in range(3,303):
    if ws.cell(column=1,row=row).value != None:
        lista_kont.append(ws.cell(column=1,row=row).value)
        if len(lista_kont) % 30 == 0:
            url += str(lista_kont).strip('['']').replace("'","").replace(' ','')
            if  ws.cell(column=2,row=1).value != None:
                url += '?date='+ws.cell(column=2,row=1).value
            else:
                url += '?date='+ str(date.today())

            response = requests.get(url)
            if response.status_code == 200:
                data = response.json()
                plik_json = data["result"]["requestId"] + '.json'
                with open(plik_json, 'w') as f:
                    lista_plikow.append(plik_json)
                    json.dump(data, f)
            else:
                print("Problem z linkiem: "+url)
            lista_kont.clear()
            url = URL_NIPY
        else:
            if ws.cell(column=1,row=row+1).value == None:
                url += str(lista_kont).strip('['']').replace("'","").replace(' ','')
                if ws.cell(column=3, row=1).value != None:
                    url += '?date=' + ws.cell(column=2, row=1).value
                else:
                    url += '?date=' + str(date.today())
                response = requests.get(url)
                if response.status_code == 200:
                    data = response.json()
                    plik_json = data["result"]["requestId"] + '.json'

                    with open(plik_json, 'w') as f:
                        lista_plikow.append(plik_json)
                        json.dump(data, f)
                else:
                    print("Problem z linkiem: "+url)

                url = URL_NIPY

for item in lista_plikow:
    with open(item) as f:
        data = json.load(f)
        for row in range(3, 303):
            nip = ws.cell(column=1, row=row).value
            for item in data['result']['subjects']:
                if str(nip) == item["nip"]:
                    ws.cell(column=4,row=row).value = item["name"]
                    ws.cell(column=5,row=row).value = item["statusVat"]
                    licznik_kont = 0
                    for account in item["accountNumbers"]:
                        if ws.cell(column=3,row=row).value == account:
                            ws.cell(column=6,row=row).value = "OK KONTO"
                        else:
                            ws.cell(column=(7 + licznik_kont), row=row).value = account
                            licznik_kont +=1

wb.save(filename='konta_bankowe.xlsx')